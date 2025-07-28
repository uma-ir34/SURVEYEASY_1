from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import PasswordChangeForm
from django.core.mail import send_mail
from django.conf import settings
from django.utils.crypto import get_random_string
from django.urls import reverse
from django.http import HttpResponse

from .models import Participant, Question, Choice, Response, Answer, AssessorRequest, Survey
from .forms import SurveyForm, ParticipantForm
from .ml_forms import MLModelForm
from .ml_pipeline import preprocess_data, train_classification, train_regression, train_knn
from .generate_report import generate_pdf_report

import os
import pandas as pd
import openpyxl


# ✅ 1️⃣ --- MACHINE LEARNING RUNNER VIEW ---
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def run_ml(request):
    result = error = preview = None
    cm_image = balance_image = scatter_image = date_image = None
    predictions_table = model_description = group_summary = None

    if request.method == 'POST':
        form = MLModelForm(request.POST, request.FILES)
        action = request.POST.get('action')

        if form.is_valid() or action:
            uploaded_file = request.FILES.get('data_file')
            if uploaded_file:
                request.session['uploaded_filename'] = uploaded_file.name
                file_path = f"/tmp/{uploaded_file.name}"
                with open(file_path, 'wb+') as dest:
                    for chunk in uploaded_file.chunks():
                        dest.write(chunk)
            else:
                file_path = f"/tmp/{request.session.get('uploaded_filename')}"
                if not os.path.exists(file_path):
                    error = "No file found. Please upload a dataset."
                    return render(request, 'ml_model_runner.html', locals())

            target_column = form.cleaned_data.get('target_column')
            model_type = request.POST.get('model_type')

            if action == 'preview':
                try:
                    df = pd.read_csv(file_path)
                    preview = df.head().to_html()
                except Exception as e:
                    error = f"Error in preview: {e}"

            elif action == 'run_model':
                try:
                    df = pd.read_csv(file_path)
                    X, y, encoders, preprocessor = preprocess_data(df, target_column)

                    if model_type == 'classification':
                        acc, report, preds_df, cm_image, balance_image, group_summary = train_classification(df, X, y, target_column)
                        result = f"✅ Classification Accuracy: {acc:.2%}\n\n📄 Report:\n{report}"
                        predictions_table = preds_df.to_html()
                        model_description = "<p>Classification explanation here.</p>"

                    elif model_type == 'regression':
                        mse, r2, preds_df, scatter_image, date_image = train_regression(df, X, y, date_col=None)
                        result = f"✅ Regression MSE: {mse:.2f} | R² Score: {r2:.2%}"
                        predictions_table = preds_df.to_html()
                        model_description = "<p>Regression explanation here.</p>"

                    elif model_type == 'knn':
                        acc, report, preds_df, cm_image, balance_image, group_summary = train_knn(df, X, y, target_column)
                        result = f"✅ KNN Classification Accuracy: {acc:.2%}\n\n📄 Report:\n{report}"
                        predictions_table = preds_df.to_html()
                        model_description = "<p>KNN explanation here.</p>"

                    else:
                        error = "Please select a valid model type."

                    if 'export_report' in request.POST:
                        pdf_buffer = generate_pdf_report(
                            title='ML Report',
                            result_text=result,
                            preds_table_html=predictions_table,
                            cm_image=cm_image,
                            balance_image=balance_image,
                            scatter_image=scatter_image,
                            date_image=date_image,
                            group_summary=group_summary,
                            model_type=model_type,
                            survey_name='Your Survey Name Here',
                            num_participants=len(df),
                            df=df
                        )
                        response = HttpResponse(
                            pdf_buffer.getvalue(),
                            content_type='application/pdf'
                        )
                        response['Content-Disposition'] = f'attachment; filename="ML_Report_{request.user.username}.pdf"'
                        return response

                except Exception as e:
                    error = f"Processing error: {e}"

        else:
            error = "Invalid form submission."

    else:
        form = MLModelForm()

    return render(request, 'ml_model_runner.html', locals())


# ✅ 2️⃣ --- EXPORT ALL RESPONSES ---
from collections import defaultdict
from django.db.models import Prefetch

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def export_all_responses(request, survey_id):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.utils import get_column_letter

    survey = get_object_or_404(Survey, id=survey_id)
    questions = survey.questions.all()
    question_texts = [q.text for q in questions]
    question_map = {q.id: q.text for q in questions}

    responses = Response.objects.filter(answers__question__survey=survey).prefetch_related(
        Prefetch('answers', queryset=Answer.objects.select_related('question', 'choice'))
    ).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Survey_{survey.id}_Responses"

    # Header row
    headers = ['Participant Name', 'Location', 'Age'] + question_texts
    ws.append(headers)

    for response in responses:
        participant = response.participant

        # Collect answers per question ID
        answers_dict = defaultdict(list)
        for ans in response.answers.all():
            if ans.question.survey_id != survey.id:
                continue
            if ans.answer_text:
                answers_dict[ans.question_id].append(ans.answer_text)
            elif ans.choice:
                answers_dict[ans.question_id].append(ans.choice.text)

        # Build row in the same order as questions
        row = [
            participant.name,
            participant.location,
            participant.age,
        ]
        for q in questions:
            combined = ", ".join(answers_dict.get(q.id, []))
            row.append(combined)

        ws.append(row)

    # Return as Excel file
    response_excel = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response_excel['Content-Disposition'] = f'attachment; filename=Survey_{survey.id}_All_Responses_Wide.xlsx'
    wb.save(response_excel)
    return response_excel





# ✅ 3️⃣ --- RESPONSE LIST + EXPORT ---
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def response_list(request):
    responses = Response.objects.select_related('participant', 'survey').order_by('survey__id', 'created_at')
    surveys = Survey.objects.all()

    if request.GET.get('export') == 'excel':
        selected_survey_id = request.GET.get('survey_id')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        if selected_survey_id:
            survey = get_object_or_404(Survey, id=selected_survey_id)
            ws = wb.create_sheet(title=survey.title[:31])
            ws.append(['Participant Name', 'Location', 'Age', 'Question', 'Answer'])
            survey_responses = responses.filter(survey=survey)
            for resp in survey_responses:
                p = resp.participant
                answers = resp.answers.select_related('question', 'choice')
                first = True
                for ans in answers:
                    ws.append([
                        p.name if first else '',
                        p.location if first else '',
                        p.age if first else '',
                        ans.question.text,
                        ans.answer_text or (ans.choice.text if ans.choice else '')
                    ])
                    first = False
        else:
            for survey in surveys:
                ws = wb.create_sheet(title=survey.title[:31])
                ws.append(['Participant Name', 'Location', 'Age', 'Question', 'Answer'])
                survey_responses = responses.filter(survey=survey)
                for resp in survey_responses:
                    p = resp.participant
                    answers = resp.answers.select_related('question', 'choice')
                    first = True
                    for ans in answers:
                        ws.append([
                            p.name if first else '',
                            p.location if first else '',
                            p.age if first else '',
                            ans.question.text,
                            ans.answer_text or (ans.choice.text if ans.choice else '')
                        ])
                        first = False

        response_excel = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response_excel['Content-Disposition'] = 'attachment; filename=Survey_Responses.xlsx'
        wb.save(response_excel)
        return response_excel

    return render(request, 'response_list.html', locals())


# ✅ 4️⃣ --- INDIVIDUAL RESPONSE EXPORT ---
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def response_detail(request, response_id):
    response = get_object_or_404(Response.objects.select_related('participant'), id=response_id)
    answers = response.answers.select_related('question', 'choice')
    participant = response.participant

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Response_{response.id}"

        ws.append(['Participant Name', participant.name])
        ws.append(['Location', participant.location])
        ws.append(['Age', participant.age])
        ws.append(['Survey', response.survey.title if response.survey else 'N/A'])
        ws.append([])
        ws.append(['Question', 'Answer'])

        for ans in answers:
            answer_value = ans.answer_text or (ans.choice.text if ans.choice else '')
            ws.append([ans.question.text, answer_value])

        response_excel = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"SurveyResponse_{participant.name}_{response.id}.xlsx"
        response_excel['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response_excel)
        return response_excel

    return render(request, 'response_detail.html', {
        'response': response,
        'answers': answers,
        'participant': participant,
    })


# ✅ 5️⃣ --- DASHBOARDS ---
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Assessor').exists())
def assessor_dashboard(request):
    return render(request, 'assessor_dashboard.html')


# ✅ SURVEY CRUD
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def create_survey(request):
    all_surveys = Survey.objects.all().order_by('-id')
    active_survey = all_surveys.first()

    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by = request.user
            survey.save()
            return redirect('add_questions', survey_id=survey.id)
    else:
        form = SurveyForm()

    return render(request, 'create_survey.html', {
        'form': form,
        'active_survey': active_survey,
        'all_surveys': all_surveys,
    })


@login_required
def survey_list(request):
    if request.user.groups.filter(name='Admin').exists():
        # Admin → see ALL
        surveys = Survey.objects.all().order_by('-id')
    else:
        # Assessor → see only their own
        surveys = Survey.objects.filter(created_by=request.user).order_by('-id')

    return render(request, 'survey_list.html', {'surveys': surveys})


@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def edit_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)

    # Enforce that assessors can only edit their own surveys
    if request.user.groups.filter(name='Assessor').exists():
        if survey.created_by != request.user:
            return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        form = SurveyForm(request.POST, instance=survey)
        if form.is_valid():
            form.save()
            return redirect('add_questions', survey_id=survey.id)
    else:
        form = SurveyForm(instance=survey)

    return render(request, 'create_survey.html', {
        'form': form,
        'active_survey': survey
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def delete_survey(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)

    if request.user.groups.filter(name='Assessor').exists():
        if survey.created_by != request.user:
            return HttpResponse("Unauthorized", status=403)

    survey.delete()
    return redirect('create_survey')



# ✅ QUESTIONS
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def add_questions(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)

    if request.user.groups.filter(name='Assessor').exists():
        if survey.created_by != request.user:
            return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        text = request.POST['text']
        qtype = request.POST['question_type']
        existing_count = Question.objects.filter(survey=survey).count()
        question = Question.objects.create(
            text=text,
            question_type=qtype,
            survey=survey,
            question_order=existing_count + 1
        )
        if qtype in ['radio', 'checkbox']:
            raw_choices = request.POST['choices']
            choice_list = [c.strip() for c in raw_choices.split(',') if c.strip()]
            for choice_text in choice_list:
                Choice.objects.create(question=question, text=choice_text)

        return redirect('add_questions', survey_id=survey.id)

    return render(request, 'add_questions.html', {
        'survey': survey
    })


@login_required
def remove_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    survey_id = question.survey.id
    question.delete()
    return redirect('add_questions', survey_id=survey_id)

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def review_questions(request, survey_id):
    survey = get_object_or_404(Survey, id=survey_id)

    if request.user.groups.filter(name='Assessor').exists():
        if survey.created_by != request.user:
            return HttpResponse("Unauthorized", status=403)

    questions = survey.questions.order_by('question_order')

    if request.method == 'POST':
        for q in questions:
            new_order = request.POST.get(f'order_{q.id}')
            if new_order:
                q.question_order = int(new_order)
                q.save()

        remove_ids = request.POST.getlist('remove')
        for qid in remove_ids:
            Question.objects.filter(id=int(qid)).delete()

        return redirect('dynamic_survey')

    return render(request, 'review_questions.html', {
        'survey': survey,
        'questions': questions
    })



# ✅ SURVEY RUN
@login_required
def dynamic_survey(request):
    survey_id = request.GET.get('survey_id')
    if survey_id:
        survey = get_object_or_404(Survey, id=survey_id)
    else:
        survey = Survey.objects.order_by('-id').first()

    questions = Question.objects.filter(survey=survey).order_by('question_order')

    if request.method == 'POST':
        participant_form = ParticipantForm(request.POST)
        if participant_form.is_valid():
            participant = participant_form.save()
            
            # ✅ IMPORTANT: Pass survey when creating the response!
            response = Response.objects.create(
                survey=survey,
                participant=participant
            )

            for question in questions:
                if question.question_type in ['text', 'number', 'date']:
                    Answer.objects.create(
                        response=response,
                        question=question,
                        answer_text=request.POST.get(f'question_{question.id}')
                    )
                elif question.question_type == 'radio':
                    selected = request.POST.get(f'question_{question.id}')
                    if selected:
                        Answer.objects.create(
                            response=response,
                            question=question,
                            choice_id=selected
                        )
                elif question.question_type == 'checkbox':
                    selected_choices = request.POST.getlist(f'question_{question.id}')
                    for cid in selected_choices:
                        Answer.objects.create(
                            response=response,
                            question=question,
                            choice_id=cid
                        )

            return redirect('review_survey', response_id=response.id)

    else:
        participant_form = ParticipantForm()

    return render(request, 'dynamic_survey.html', {
        'survey': survey,
        'questions': questions,
        'participant_form': participant_form
    })


@login_required
def review_survey(request, response_id):
    response = get_object_or_404(Response, id=response_id)
    answers = response.answers.select_related('question', 'choice')
    participant = response.participant

    return render(request, 'review_survey.html', {
        'participant': participant,
        'answers': answers,
        'survey': response.participant.responses.first().participant.responses.first().participant.survey if hasattr(response.participant, 'survey') else None
    })


# ✅ AUTH
def user_login(request):
    error = False
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if role == 'Admin' and user.groups.filter(name='Admin').exists():
                login(request, user)
                return redirect('admin_dashboard')
            elif role == 'Assessor' and user.groups.filter(name='Assessor').exists():
                login(request, user)
                return redirect('assessor_dashboard')
            else:
                error = "Role mismatch or not allowed."
        else:
            error = "Invalid username or password."

    return render(request, 'login.html', {'error': error})


def user_logout(request):
    logout(request)
    return redirect('login')

@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return render(request, 'password_changed.html')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'change_password.html', {'form': form})


# ✅ REGISTRATION & APPROVAL
def register_assessor(request):
    success = False
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        reason = request.POST['reason']

        req = AssessorRequest.objects.create(name=name, email=email, reason=reason)
        approve_url = f"{request.scheme}://{request.get_host()}/approve/{req.token}/"
        decline_url = f"{request.scheme}://{request.get_host()}/decline/{req.token}/"

        send_mail(
            'New Assessor Registration Request',
            f"Name: {name}\nEmail: {email}\nReason: {reason}\n\nApprove: {approve_url}\nDecline: {decline_url}",
            settings.DEFAULT_FROM_EMAIL,
            [settings.ADMIN_EMAIL]
        )

        success = True

    return render(request, 'register.html', {'success': success})


def approve_assessor(request, token):
    try:
        req = AssessorRequest.objects.get(token=token, approved=False, declined=False)
    except AssessorRequest.DoesNotExist:
        return HttpResponse("Invalid or already processed.", status=400)

    username = req.email
    password = get_random_string(10)

    user = User.objects.create_user(username=username, email=req.email, password=password)
    assessor_group, _ = Group.objects.get_or_create(name='Assessor')
    user.groups.add(assessor_group)

    req.approved = True
    req.save()

    send_mail(
        "Your Assessor Access is Approved",
        f"Username (your email): {username}\nPassword: {password}\nLogin: {request.scheme}://{request.get_host()}/\nPlease change your password immediately.",
        settings.DEFAULT_FROM_EMAIL,
        [req.email]
    )
    return HttpResponse("Assessor Approved and credentials sent.")


def decline_assessor(request, token):
    try:
        req = AssessorRequest.objects.get(token=token, approved=False, declined=False)
    except AssessorRequest.DoesNotExist:
        return HttpResponse("Invalid or already processed.", status=400)
    req.declined = True
    req.save()
    return HttpResponse("Assessor Request Declined.")
def thank_you(request):
    return render(request, 'thank_you.html')


from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
import openpyxl

from .models import Response


from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
import openpyxl

from .models import Response


@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Admin', 'Assessor']).exists())
def response_detail(request, response_id):
    response = get_object_or_404(Response.objects.select_related('participant'), id=response_id)
    answers = response.answers.select_related('question', 'choice')
    participant = response.participant

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Response_{response.id}"

        # Participant info
        ws.append(['Participant Name', participant.name])
        ws.append(['Location', participant.location])
        ws.append(['Age', participant.age])
        ws.append(['Survey', response.survey.title if response.survey else 'N/A'])
        ws.append([])
        ws.append(['Question', 'Answer'])

        # Answers
        for ans in answers:
            answer_value = ans.answer_text or (ans.choice.text if ans.choice else '')
            ws.append([ans.question.text, answer_value])

        response_excel = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"SurveyResponse_{participant.name}_{response.id}.xlsx"
        response_excel['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response_excel)
        return response_excel

    return render(request, 'response_detail.html', {
        'response': response,
        'answers': answers,
        'participant': participant,
    })



